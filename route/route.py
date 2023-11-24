import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

font      = Font(color="FFFFFF")
fill      = PatternFill(fill_type='solid', start_color='00000000')
alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
border    = Border(left=Side(border_style='thin', color='00000000'), right=Side(border_style='thin', color='00000000'),
                   top=Side(border_style='thin', color='00000000'),bottom=Side(border_style='thin', color='00000000'))

xp_pattern  = r"x(\d+) y(\d+)"
port_pattern = r"(?P<direction>tx|rx)_(?P<port>\d+) pps:    (?P<pps>\d+)|            (?P<cnt>\d+) cnt"
que_pattern = r"port:(?P<port>\d+)  dir:(?P<dir>up|east|west|north|south|inter)"
new_line = False
row_num  = 4
col_num  = 3
step     = 5

xy_coordinates = []
port_coordinates = []

wb = openpyxl.Workbook()
ws = wb.active

def get_xp():
    count = 0
    with open("log.log", "r") as f:
        for line in f:
            match = re.search(xp_pattern, line)
            if match:
                port_coordinates.clear()
                x = match.group(1)
                y = match.group(2)
            
                if int(x) == 0:
                    step    = 5
                    col_num = 3
                    
                    ws.insert_rows(1, row_num + 1)

                    step = step * count
                    count = count + 1

                    col_num = col_num + 1
                    xy_coordinates.append((row_num + step, col_num))
                    ws.cell(row=row_num, column=col_num).value = int(x)

                    col_num = col_num + 1
                    xy_coordinates.append((row_num + step, col_num))
                    ws.cell(row=row_num, column=col_num).value = int(y)
                else:
                    col_num = col_num + 4
                    xy_coordinates.append((row_num + step, col_num))
                    ws.cell(row=row_num, column=col_num).value = int(x)

                    col_num = col_num + 1
                    xy_coordinates.append((row_num + step, col_num))
                    ws.cell(row=row_num, column=col_num).value = int(y)
            
            match = re.search(port_pattern, line)
            if match:
                result = match.groups()
                dir = result[0]
                port = int(result[1])
                pps = int(result[2])
                #print(dir, port, pps)

                port_coordinates.append((dir, port, pps))
            
            match = re.search(que_pattern, line)
            if match:
                result = match.groups()
                port = int(result[0])
                dir = result[1]
                
                if dir == "east":
                    ws.cell(row=row_num, column=col_num + 1).value = 0
                    ws.cell(row=row_num + 1, column=col_num + 1).value = 0
                    ws.cell(row=row_num, column=col_num + 1).alignment = alignment
                    ws.cell(row=row_num, column=col_num + 1).border = border
                    ws.cell(row=row_num, column=col_num + 2).value = "→"
                    ws.cell(row=row_num, column=col_num + 2).alignment = alignment
                    ws.cell(row=row_num, column=col_num + 2).fill = PatternFill(fill_type='solid', start_color='FFC0CB')
                    ws.cell(row=row_num + 1, column=col_num + 1).alignment = alignment
                    ws.cell(row=row_num + 1, column=col_num + 1).border = border
                
                if dir == "west":
                    ws.cell(row=row_num + 1, column=col_num - 2).value = 0
                    ws.cell(row=row_num, column=col_num - 2).value = 0
                    ws.cell(row=row_num + 1, column=col_num - 2).alignment = alignment
                    ws.cell(row=row_num + 1, column=col_num - 2).border = border
                    ws.cell(row=row_num + 1, column=col_num - 3).value = "←"
                    ws.cell(row=row_num + 1, column=col_num - 3).alignment = alignment
                    ws.cell(row=row_num + 1, column=col_num - 3).fill = PatternFill(fill_type='solid', start_color='87CEEB')
                    ws.cell(row=row_num, column=col_num - 2).alignment = alignment
                    ws.cell(row=row_num, column=col_num - 2).border = border
                
                if dir == "south":
                    ws.cell(row=row_num + 2, column=col_num).value = 0
                    ws.cell(row=row_num + 2, column=col_num  - 1).value = 0
                    ws.cell(row=row_num + 2, column=col_num).alignment = alignment
                    ws.cell(row=row_num + 2, column=col_num).border = border
                    ws.cell(row=row_num + 3, column=col_num).value = "↓"
                    ws.cell(row=row_num + 3, column=col_num).alignment = alignment
                    ws.cell(row=row_num + 3, column=col_num).fill = PatternFill(fill_type='solid', start_color='87CEEB')
                    ws.cell(row=row_num + 2, column=col_num  - 1).alignment = alignment
                    ws.cell(row=row_num + 2, column=col_num  - 1).border = border
                
                if dir == "north":
                    ws.cell(row=row_num - 1, column=col_num - 1).value = 0
                    ws.cell(row=row_num - 1, column=col_num).value = 0
                    ws.cell(row=row_num - 1, column=col_num - 1).alignment = alignment
                    ws.cell(row=row_num - 1, column=col_num - 1).border = border
                    ws.cell(row=row_num - 2, column=col_num - 1).value = "↑"
                    ws.cell(row=row_num - 2, column=col_num - 1).alignment = alignment
                    ws.cell(row=row_num - 2, column=col_num - 1).fill = PatternFill(fill_type='solid', start_color='FFC0CB')
                    ws.cell(row=row_num - 1, column=col_num).alignment = alignment
                    ws.cell(row=row_num - 1, column=col_num).border = border

                for p in port_coordinates:
                    if p[1] == port:
                        if dir == "east" and p[0] == "tx":
                            ws.cell(row=row_num, column=col_num + 1).value = p[2]
                        if dir == "east" and p[0] == "rx":
                            ws.cell(row=row_num + 1, column=col_num + 1).value = p[2]
                        if dir == "west" and p[0] == "tx":
                            ws.cell(row=row_num + 1, column=col_num - 2).value = p[2]
                        if dir == "west" and p[0] == "rx":
                            ws.cell(row=row_num, column=col_num - 2).value = p[2]
                        if dir == "south" and p[0] == "tx":
                            ws.cell(row=row_num + 2, column=col_num).value = p[2]
                        if dir == "south" and p[0] == "rx":
                            ws.cell(row=row_num + 2, column=col_num  - 1).value = p[2]
                        if dir == "north" and p[0] == "tx":
                            ws.cell(row=row_num - 1, column=col_num - 1).value = p[2]
                        if dir == "north" and p[0] == "rx":
                            ws.cell(row=row_num - 1, column=col_num).value = p[2]
                        
                        print(p[0], p[1], dir, p[2])
                    

def pait_xp():
    for row in xy_coordinates:
        x = row[0]
        y = row[1]

        cell1 = ws.cell(row=x, column=y)
        cell2 = ws.cell(row=x+1, column=y)
        ws.merge_cells(cell1.coordinate + ":" + cell2.coordinate)

        ws.cell(row=x, column=y).alignment = alignment
        ws.cell(row=x, column=y).fill      = fill
        ws.cell(row=x, column=y).border    = border
        ws.cell(row=x, column=y).font      = font

        #ws.cell(row=x, column=y).value = "↑"
        #ws.row_dimensions[row_num + 3].height=100
        #ws.column_dimensions[col_num - 3].width=50


def main():
    get_xp()
    pait_xp()
    wb.save("route.xlsx")

main()