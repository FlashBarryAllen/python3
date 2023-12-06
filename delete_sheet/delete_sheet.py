import openpyxl

wb = openpyxl.Workbook()
wb.create_sheet()
wb.create_sheet()
wb.create_sheet()
wb.save('testdel.xlsx')

workbook=openpyxl.load_workbook('testdel.xlsx')
workbook.get_sheet_names()
std=workbook.get_sheet_by_name('Sheet2')
workbook.remove_sheet(std)
workbook.get_sheet_names()
workbook.save('testdel.xlsx')
